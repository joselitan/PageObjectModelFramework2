import openpyxl


def get_data(sheetName):

    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    main_list = []

    for i in range(2, totalrows+1):
        data_list = []
        for j in range(1, totalcols+1):
            data = sheet.cell(row=i, column=j).value
            data_list.insert(j, data)
        main_list.insert(i, data_list)
    return main_list
